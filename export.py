"""
Export and Download module for Campus Assets system.
Handles CSV/Excel exports, department summaries, and inventory reports.
"""
import io
import csv
import json
import os  # Added os import
from datetime import datetime, date, timedelta
from flask import Blueprint, request, jsonify, Response, make_response  # make_response imported here
from typing import Dict, List, Optional, Any
import logging
import csv
import io
import json
import requests
from bson import ObjectId
from database import get_db
import pandas as pd

from auth import require_auth, require_role
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference
import tempfile
import zipfile
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4


# Configure logging
logger = logging.getLogger(__name__)

# Create Blueprint
export_bp = Blueprint('export', __name__)

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def format_date_for_export(date_obj):
    """Format date objects for export."""
    if date_obj is None:
        return ""
    
    if isinstance(date_obj, str):
        try:
            # Try to parse string date
            date_obj = datetime.fromisoformat(date_obj.replace('Z', '+00:00'))
        except:
            return date_obj
    
    if isinstance(date_obj, (datetime, date)):
        return date_obj.strftime('%Y-%m-%d')
    
    return str(date_obj)

def prepare_resource_data(resources):
    """Prepare resource data for export with proper date formatting."""
    export_data = []
    
    for resource in resources:
        export_row = {
            'sl_no': resource.get('sl_no', ''),
            'device_name': resource.get('device_name', ''),
            'quantity': resource.get('quantity', 0),
            'description': resource.get('description', ''),
            'procurement_date': format_date_for_export(resource.get('procurement_date')),
            'location': resource.get('location', ''),
            'cost': resource.get('cost', 0),
            'total_value': resource.get('cost', 0) * resource.get('quantity', 0),
            'department': resource.get('department', ''),
            'created_at': format_date_for_export(resource.get('created_at')),
            'updated_at': format_date_for_export(resource.get('updated_at'))
        }
        export_data.append(export_row)
    
    return export_data

def apply_excel_styling(worksheet, title="Campus Assets Report"):
    """Apply professional styling to Excel worksheet."""
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header styling
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border

def generate_summary_statistics(resources):
    """Generate summary statistics for resources."""
    if not resources:
        return {}
    
    df = pd.DataFrame(prepare_resource_data(resources))
    
    summary = {
        'total_items': len(resources),
        'total_quantity': df['quantity'].sum(),
        'total_value': df['total_value'].sum(),
        'average_cost_per_item': df['cost'].mean(),
        'most_expensive_item': df['cost'].max(),
        'least_expensive_item': df['cost'].min(),
        'unique_departments': df['department'].nunique(),
        'unique_locations': df['location'].nunique(),
        'unique_device_types': df['device_name'].nunique(),
        'department_breakdown': df.groupby('department')['quantity'].sum().to_dict(),
        'location_breakdown': df.groupby('location')['quantity'].sum().to_dict(),
        'device_type_breakdown': df.groupby('device_name')['quantity'].sum().to_dict(),
        'cost_by_department': df.groupby('department')['total_value'].sum().to_dict(),
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    return summary

# ============================================================================
# BASIC EXPORT ENDPOINTS
# ============================================================================

@export_bp.route('/csv', methods=['GET'])
@require_auth
def export_csv():
    """Export resources to CSV format."""
    try:
        db = get_db()
        
        # Get query parameters for filtering
        department = request.args.get('department')
        location = request.args.get('location')
        device_type = request.args.get('device_type')
        search = request.args.get('search')
        
        # Build query filter
        query_filter = {}
        
        if department:
            query_filter['department'] = department
        if location:
            query_filter['location'] = location
        if device_type:
            query_filter['device_name'] = {'$regex': device_type, '$options': 'i'}
        if search:
            query_filter['$or'] = [
                {'device_name': {'$regex': search, '$options': 'i'}},
                {'description': {'$regex': search, '$options': 'i'}},
                {'location': {'$regex': search, '$options': 'i'}}
            ]
        
        # Get resources
        resources = list(db.resources.find(query_filter).sort('sl_no', 1))
        
        # Prepare data for export
        export_data = prepare_resource_data(resources)
        
        # Create CSV
        output = io.StringIO()
        fieldnames = ['sl_no', 'device_name', 'quantity', 'description', 'procurement_date', 
                     'location', 'cost', 'total_value', 'department', 'created_at', 'updated_at']
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        
        writer.writeheader()
        writer.writerows(export_data)
        
        csv_content = output.getvalue()
        output.close()
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'campus_assets_export_{timestamp}.csv'
        
        return Response(
            csv_content,
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
        
    except Exception as e:
        logger.error(f"Error exporting CSV: {e}")
        return jsonify({'error': 'Failed to export CSV'}), 500

@export_bp.route('/excel', methods=['GET'])
@require_auth
def export_excel():
    """Export resources to Excel format with multiple sheets."""
    try:
        db = get_db()
        
        # Get query parameters for filtering
        department = request.args.get('department')
        location = request.args.get('location')
        device_type = request.args.get('device_type')
        search = request.args.get('search')
        
        # Build query filter
        query_filter = {}
        
        if department:
            query_filter['department'] = department
        if location:
            query_filter['location'] = location
        if device_type:
            query_filter['device_name'] = {'$regex': device_type, '$options': 'i'}
        if search:
            query_filter['$or'] = [
                {'device_name': {'$regex': search, '$options': 'i'}},
                {'description': {'$regex': search, '$options': 'i'}},
                {'location': {'$regex': search, '$options': 'i'}}
            ]
        
        # Get resources
        resources = list(db.resources.find(query_filter).sort('sl_no', 1))
        
        # Prepare data for export
        export_data = prepare_resource_data(resources)
        summary_stats = generate_summary_statistics(resources)
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main data sheet
            df = pd.DataFrame(export_data)
            df.to_excel(writer, sheet_name='Resources', index=False)
            
            # Summary sheet
            if resources:
                summary_data = {
                    'Metric': [
                        'Total Items',
                        'Total Quantity',
                        'Total Value (₹)',
                        'Average Cost per Item (₹)',
                        'Most Expensive Item (₹)',
                        'Least Expensive Item (₹)',
                        'Unique Departments',
                        'Unique Locations',
                        'Unique Device Types',
                        'Report Generated'
                    ],
                    'Value': [
                        summary_stats['total_items'],
                        summary_stats['total_quantity'],
                        f"₹{summary_stats['total_value']:,.2f}",
                        f"₹{summary_stats['average_cost_per_item']:,.2f}",
                        f"₹{summary_stats['most_expensive_item']:,.2f}",
                        f"₹{summary_stats['least_expensive_item']:,.2f}",
                        summary_stats['unique_departments'],
                        summary_stats['unique_locations'],
                        summary_stats['unique_device_types'],
                        summary_stats['generated_at']
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Department breakdown sheet
                dept_data = []
                for dept, quantity in summary_stats['department_breakdown'].items():
                    cost = summary_stats['cost_by_department'].get(dept, 0)
                    dept_data.append({
                        'Department': dept,
                        'Total Quantity': quantity,
                        'Total Value (₹)': cost,
                        'Average Cost per Item (₹)': cost / quantity if quantity > 0 else 0
                    })
                
                dept_df = pd.DataFrame(dept_data)
                dept_df.to_excel(writer, sheet_name='Department Breakdown', index=False)
            
            # Apply styling to all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                apply_excel_styling(worksheet)
        
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'campus_assets_export_{timestamp}.xlsx'
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
        
    except Exception as e:
        logger.error(f"Error exporting Excel: {e}")
        return jsonify({'error': 'Failed to export Excel'}), 500

# ============================================================================
# DEPARTMENT SUMMARY REPORTS
# ============================================================================

@export_bp.route('/department-summary/<department_name>', methods=['GET'])
@require_auth
def export_department_summary(department_name):
    """Generate comprehensive department summary report."""
    try:
        db = get_db()
        format_type = request.args.get('format', 'excel').lower()
        
        # Get department resources
        resources = list(db.resources.find({'department': department_name}).sort('sl_no', 1))
        
        if not resources:
            return jsonify({'error': f'No resources found for department: {department_name}'}), 404
        
        # Generate comprehensive analytics
        department_analytics = generate_department_analytics(department_name, resources)
        
        if format_type == 'excel':
            return generate_department_excel_report(department_name, resources, department_analytics)
        else:
            return generate_department_csv_report(department_name, resources, department_analytics)
            
    except Exception as e:
        logger.error(f"Error generating department summary: {e}")
        return jsonify({'error': 'Failed to generate department summary'}), 500

def generate_department_analytics(department_name, resources):
    """Generate comprehensive department analytics."""
    df = pd.DataFrame(prepare_resource_data(resources))
    
    analytics = {
        'department_name': department_name,
        'total_resources': len(resources),
        'total_quantity': df['quantity'].sum(),
        'total_value': df['total_value'].sum(),
        'average_cost_per_item': df['cost'].mean(),
        'most_expensive_item': {
            'name': df.loc[df['cost'].idxmax(), 'device_name'],
            'cost': df['cost'].max(),
            'location': df.loc[df['cost'].idxmax(), 'location']
        },
        'least_expensive_item': {
            'name': df.loc[df['cost'].idxmin(), 'device_name'],
            'cost': df['cost'].min(),
            'location': df.loc[df['cost'].idxmin(), 'location']
        },
        'unique_locations': df['location'].nunique(),
        'unique_device_types': df['device_name'].nunique(),
        'location_breakdown': df.groupby('location').agg({
            'quantity': 'sum',
            'total_value': 'sum',
            'device_name': 'nunique'
        }).to_dict(),
        'device_type_breakdown': df.groupby('device_name').agg({
            'quantity': 'sum',
            'total_value': 'sum',
            'cost': 'mean'
        }).to_dict(),
        'cost_distribution': {
            'high_value_items': len(df[df['cost'] > 100000]),
            'medium_value_items': len(df[(df['cost'] >= 50000) & (df['cost'] <= 100000)]),
            'low_value_items': len(df[df['cost'] < 50000])
        },
        'procurement_timeline': df.groupby('procurement_date')['quantity'].sum().to_dict(),
        'recent_additions': len(df[pd.to_datetime(df['procurement_date']) > (datetime.now() - timedelta(days=90))]),
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    return analytics

def generate_department_excel_report(department_name, resources, analytics):
    """Generate comprehensive department Excel report."""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Main resources sheet
        df = pd.DataFrame(prepare_resource_data(resources))
        df.to_excel(writer, sheet_name='Resources', index=False)
        
        # Department overview sheet
        overview_data = {
            'Metric': [
                'Department Name',
                'Total Resources',
                'Total Quantity',
                'Total Value (₹)',
                'Average Cost per Item (₹)',
                'Unique Locations',
                'Unique Device Types',
                'High Value Items (>₹1L)',
                'Medium Value Items (₹50K-₹1L)',
                'Low Value Items (<₹50K)',
                'Recent Additions (90 days)',
                'Report Generated'
            ],
            'Value': [
                analytics['department_name'],
                analytics['total_resources'],
                analytics['total_quantity'],
                f"₹{analytics['total_value']:,.2f}",
                f"₹{analytics['average_cost_per_item']:,.2f}",
                analytics['unique_locations'],
                analytics['unique_device_types'],
                analytics['cost_distribution']['high_value_items'],
                analytics['cost_distribution']['medium_value_items'],
                analytics['cost_distribution']['low_value_items'],
                analytics['recent_additions'],
                analytics['generated_at']
            ]
        }
        overview_df = pd.DataFrame(overview_data)
        overview_df.to_excel(writer, sheet_name='Department Overview', index=False)
        
        # Location breakdown sheet
        location_data = []
        for location in analytics['location_breakdown']['quantity']:
            location_data.append({
                'Location': location,
                'Total Quantity': analytics['location_breakdown']['quantity'][location],
                'Total Value (₹)': analytics['location_breakdown']['total_value'][location],
                'Device Types': analytics['location_breakdown']['device_name'][location]
            })
        
        location_df = pd.DataFrame(location_data)
        location_df.to_excel(writer, sheet_name='Location Breakdown', index=False)
        
        # Device type breakdown sheet
        device_data = []
        for device in analytics['device_type_breakdown']['quantity']:
            device_data.append({
                'Device Type': device,
                'Total Quantity': analytics['device_type_breakdown']['quantity'][device],
                'Total Value (₹)': analytics['device_type_breakdown']['total_value'][device],
                'Average Cost (₹)': analytics['device_type_breakdown']['cost'][device]
            })
        
        device_df = pd.DataFrame(device_data)
        device_df.to_excel(writer, sheet_name='Device Breakdown', index=False)
        
        # Top items sheet (most expensive)
        top_items = df.nlargest(20, 'cost')[['device_name', 'quantity', 'cost', 'total_value', 'location']]
        top_items.to_excel(writer, sheet_name='Top Expensive Items', index=False)
        
        # Apply styling to all sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            apply_excel_styling(worksheet, f"{department_name} - Campus Assets Report")
    
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{department_name.replace(" ", "_")}_summary_{timestamp}.xlsx'
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

def generate_department_csv_report(department_name, resources, analytics):
    """Generate department CSV report."""
    export_data = prepare_resource_data(resources)
    
    output = io.StringIO()
    fieldnames = ['sl_no', 'device_name', 'quantity', 'description', 'procurement_date', 
                 'location', 'cost', 'total_value', 'department']
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    
    # Add summary header
    output.write(f"# Department Summary Report: {department_name}\n")
    output.write(f"# Generated: {analytics['generated_at']}\n")
    output.write(f"# Total Resources: {analytics['total_resources']}\n")
    output.write(f"# Total Value: ₹{analytics['total_value']:,.2f}\n")
    output.write(f"# Unique Locations: {analytics['unique_locations']}\n")
    output.write("#\n")
    
    writer.writeheader()
    writer.writerows(export_data)
    
    csv_content = output.getvalue()
    output.close()
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{department_name.replace(" ", "_")}_summary_{timestamp}.csv'
    
    return Response(
        csv_content,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

# ============================================================================
# INVENTORY REPORTS
# ============================================================================

@export_bp.route('/inventory-report', methods=['GET'])
@require_auth
def generate_inventory_report():
    """Generate comprehensive inventory report."""
    try:
        report_type = request.args.get('type', 'full')  # full, summary, critical, location
        format_type = request.args.get('format', 'excel').lower()
        
        if report_type == 'full':
            return generate_full_inventory_report(format_type)
        elif report_type == 'summary':
            return generate_summary_inventory_report(format_type)
        elif report_type == 'critical':
            return generate_critical_inventory_report(format_type)
        elif report_type == 'location':
            return generate_location_inventory_report(format_type)
        else:
            return jsonify({'error': 'Invalid report type'}), 400
            
    except Exception as e:
        logger.error(f"Error generating inventory report: {e}")
        return jsonify({'error': 'Failed to generate inventory report'}), 500

def generate_full_inventory_report(format_type):
    """Generate full inventory report with all details."""
    db = get_db()
    
    # Get all resources
    resources = list(db.resources.find({}).sort('department', 1).sort('location', 1))
    
    # Generate comprehensive analytics
    inventory_analytics = generate_inventory_analytics(resources)
    
    if format_type == 'excel':
        return create_full_inventory_excel(resources, inventory_analytics)
    else:
        return create_full_inventory_csv(resources, inventory_analytics)

def generate_inventory_analytics(resources):
    """Generate comprehensive inventory analytics."""
    df = pd.DataFrame(prepare_resource_data(resources))
    
    current_date = datetime.now()
    
    analytics = {
        'total_items': len(resources),
        'total_quantity': df['quantity'].sum(),
        'total_value': df['total_value'].sum(),
        'departments': {
            'count': df['department'].nunique(),
            'breakdown': df.groupby('department').agg({
                'quantity': 'sum',
                'total_value': 'sum',
                'device_name': 'nunique'
            }).to_dict()
        },
        'locations': {
            'count': df['location'].nunique(),
            'breakdown': df.groupby('location').agg({
                'quantity': 'sum',
                'total_value': 'sum'
            }).to_dict()
        },
        'device_types': {
            'count': df['device_name'].nunique(),
            'breakdown': df.groupby('device_name').agg({
                'quantity': 'sum',
                'total_value': 'sum',
                'cost': 'mean'
            }).to_dict()
        },
        'value_distribution': {
            'high_value': len(df[df['cost'] > 100000]),
            'medium_value': len(df[(df['cost'] >= 50000) & (df['cost'] <= 100000)]),
            'low_value': len(df[df['cost'] < 50000])
        },
        'age_analysis': analyze_equipment_age(df),
        'utilization_metrics': analyze_resource_utilization(df),
        'generated_at': current_date.strftime('%Y-%m-%d %H:%M:%S')
    }
    
    return analytics

def analyze_equipment_age(df):
    """Analyze equipment age distribution."""
    current_date = datetime.now()
    
    age_data = []
    for _, row in df.iterrows():
        try:
            proc_date = pd.to_datetime(row['procurement_date'])
            age_days = (current_date - proc_date).days
            age_years = age_days / 365.25
            
            if age_years < 1:
                category = 'New (< 1 year)'
            elif age_years < 3:
                category = 'Recent (1-3 years)'
            elif age_years < 5:
                category = 'Mature (3-5 years)'
            else:
                category = 'Old (> 5 years)'
            
            age_data.append(category)
        except:
            age_data.append('Unknown')
    
    age_df = pd.DataFrame({'age_category': age_data, 'quantity': df['quantity']})
    age_breakdown = age_df.groupby('age_category')['quantity'].sum().to_dict()
    
    return age_breakdown

def analyze_resource_utilization(df):
    """Analyze resource utilization patterns."""
    utilization = {
        'avg_resources_per_location': df.groupby('location')['quantity'].sum().mean(),
        'max_resources_location': df.groupby('location')['quantity'].sum().max(),
        'min_resources_location': df.groupby('location')['quantity'].sum().min(),
        'most_common_device': df['device_name'].value_counts().index[0] if not df.empty else 'N/A',
        'device_diversity_score': df['device_name'].nunique() / len(df) if not df.empty else 0
    }
    
    return utilization

def create_full_inventory_excel(resources, analytics):
    """Create comprehensive Excel inventory report."""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Main inventory sheet
        df = pd.DataFrame(prepare_resource_data(resources))
        df.to_excel(writer, sheet_name='Full Inventory', index=False)
        
        # Executive summary sheet
        summary_data = {
            'Metric': [
                'Total Items',
                'Total Quantity',
                'Total Asset Value (₹)',
                'Number of Departments',
                'Number of Locations',
                'Number of Device Types',
                'High Value Items (>₹1L)',
                'Medium Value Items (₹50K-₹1L)',
                'Low Value Items (<₹50K)',
                'Average Resources per Location',
                'Most Common Device Type',
                'Device Diversity Score',
                'Report Generated'
            ],
            'Value': [
                analytics['total_items'],
                analytics['total_quantity'],
                f"₹{analytics['total_value']:,.2f}",
                analytics['departments']['count'],
                analytics['locations']['count'],
                analytics['device_types']['count'],
                analytics['value_distribution']['high_value'],
                analytics['value_distribution']['medium_value'],
                analytics['value_distribution']['low_value'],
                f"{analytics['utilization_metrics']['avg_resources_per_location']:.1f}",
                analytics['utilization_metrics']['most_common_device'],
                f"{analytics['utilization_metrics']['device_diversity_score']:.3f}",
                analytics['generated_at']
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
        
        # Department analysis sheet
        dept_data = []
        for dept in analytics['departments']['breakdown']['quantity']:
            dept_data.append({
                'Department': dept,
                'Total Quantity': analytics['departments']['breakdown']['quantity'][dept],
                'Total Value (₹)': analytics['departments']['breakdown']['total_value'][dept],
                'Device Types': analytics['departments']['breakdown']['device_name'][dept],
                'Value per Item (₹)': analytics['departments']['breakdown']['total_value'][dept] / analytics['departments']['breakdown']['quantity'][dept]
            })
        
        dept_df = pd.DataFrame(dept_data).sort_values('Total Value (₹)', ascending=False)
        dept_df.to_excel(writer, sheet_name='Department Analysis', index=False)
        
        # Location analysis sheet
        location_data = []
        for location in analytics['locations']['breakdown']['quantity']:
            location_data.append({
                'Location': location,
                'Total Quantity': analytics['locations']['breakdown']['quantity'][location],
                'Total Value (₹)': analytics['locations']['breakdown']['total_value'][location]
            })
        
        location_df = pd.DataFrame(location_data).sort_values('Total Value (₹)', ascending=False)
        location_df.to_excel(writer, sheet_name='Location Analysis', index=False)
        
        # Device type analysis sheet
        device_data = []
        for device in analytics['device_types']['breakdown']['quantity']:
            device_data.append({
                'Device Type': device,
                'Total Quantity': analytics['device_types']['breakdown']['quantity'][device],
                'Total Value (₹)': analytics['device_types']['breakdown']['total_value'][device],
                'Average Cost (₹)': analytics['device_types']['breakdown']['cost'][device]
            })
        
        device_df = pd.DataFrame(device_data).sort_values('Total Value (₹)', ascending=False)
        device_df.to_excel(writer, sheet_name='Device Analysis', index=False)
        
        # Age analysis sheet
        age_data = []
        for category, quantity in analytics['age_analysis'].items():
            age_data.append({
                'Age Category': category,
                'Quantity': quantity,
                'Percentage': (quantity / analytics['total_quantity'] * 100) if analytics['total_quantity'] > 0 else 0
            })
        
        age_df = pd.DataFrame(age_data)
        age_df.to_excel(writer, sheet_name='Age Analysis', index=False)
        
        # High value items sheet
        high_value_items = df[df['cost'] > 100000].sort_values('cost', ascending=False)
        if not high_value_items.empty:
            high_value_items.to_excel(writer, sheet_name='High Value Items', index=False)
        
        # Apply styling to all sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            apply_excel_styling(worksheet, "Campus Assets - Comprehensive Inventory Report")
    
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'comprehensive_inventory_report_{timestamp}.xlsx'
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

def create_full_inventory_csv(resources, analytics):
    """Create comprehensive CSV inventory report."""
    export_data = prepare_resource_data(resources)
    
    output = io.StringIO()
    
    # Add comprehensive header
    output.write("# CAMPUS ASSETS - COMPREHENSIVE INVENTORY REPORT\n")
    output.write(f"# Generated: {analytics['generated_at']}\n")
    output.write("#\n")
    output.write("# EXECUTIVE SUMMARY\n")
    output.write(f"# Total Items: {analytics['total_items']}\n")
    output.write(f"# Total Quantity: {analytics['total_quantity']}\n")
    output.write(f"# Total Value: ₹{analytics['total_value']:,.2f}\n")
    output.write(f"# Departments: {analytics['departments']['count']}\n")
    output.write(f"# Locations: {analytics['locations']['count']}\n")
    output.write(f"# Device Types: {analytics['device_types']['count']}\n")
    output.write("#\n")
    
    fieldnames = ['sl_no', 'device_name', 'quantity', 'description', 'procurement_date', 
                 'location', 'cost', 'total_value', 'department']
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    
    writer.writeheader()
    writer.writerows(export_data)
    
    csv_content = output.getvalue()
    output.close()
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'comprehensive_inventory_report_{timestamp}.csv'
    
    return Response(
        csv_content,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

# ============================================================================
# SPECIALIZED REPORTS
# ============================================================================

@export_bp.route('/critical-items', methods=['GET'])
@require_auth
def export_critical_items():
    """Export critical/high-value items report."""
    try:
        db = get_db()
        format_type = request.args.get('format', 'excel').lower()
        threshold = float(request.args.get('threshold', 100000))  # Default ₹1L
        
        # Get high-value items
        critical_items = list(db.resources.find({'cost': {'$gte': threshold}}).sort('cost', -1))
        
        if not critical_items:
            return jsonify({'error': f'No critical items found above ₹{threshold:,.2f}'}), 404
        
        if format_type == 'excel':
            return create_critical_items_excel(critical_items, threshold)
        else:
            return create_critical_items_csv(critical_items, threshold)
            
    except Exception as e:
        logger.error(f"Error exporting critical items: {e}")
        return jsonify({'error': 'Failed to export critical items'}), 500

def create_critical_items_excel(critical_items, threshold):
    """Create Excel report for critical items."""
    output = io.BytesIO()
    
    df = pd.DataFrame(prepare_resource_data(critical_items))
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Critical items sheet
        df.to_excel(writer, sheet_name='Critical Items', index=False)
        
        # Summary sheet
        summary_data = {
            'Metric': [
                'Threshold Value (₹)',
                'Total Critical Items',
                'Total Value (₹)',
                'Average Cost (₹)',
                'Most Expensive Item',
                'Most Expensive Cost (₹)',
                'Departments with Critical Items',
                'Locations with Critical Items'
            ],
            'Value': [
                f"₹{threshold:,.2f}",
                len(critical_items),
                f"₹{df['total_value'].sum():,.2f}",
                f"₹{df['cost'].mean():,.2f}",
                df.loc[df['cost'].idxmax(), 'device_name'],
                f"₹{df['cost'].max():,.2f}",
                df['department'].nunique(),
                df['location'].nunique()
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Apply styling
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            apply_excel_styling(worksheet, f"Critical Items Report (>₹{threshold:,.2f})")
    
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'critical_items_report_{timestamp}.xlsx'
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

def create_critical_items_csv(critical_items, threshold):
    """Create CSV report for critical items."""
    export_data = prepare_resource_data(critical_items)
    
    output = io.StringIO()
    
    # Add header
    output.write(f"# CRITICAL ITEMS REPORT (Threshold: ₹{threshold:,.2f})\n")
    output.write(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    output.write(f"# Total Critical Items: {len(critical_items)}\n")
    output.write("#\n")
    
    fieldnames = ['sl_no', 'device_name', 'quantity', 'description', 'procurement_date', 
                 'location', 'cost', 'total_value', 'department']
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    
    writer.writeheader()
    writer.writerows(export_data)
    
    csv_content = output.getvalue()
    output.close()
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'critical_items_report_{timestamp}.csv'
    
    return Response(
        csv_content,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

@export_bp.route('/location/<location_name>', methods=['GET'])
@require_auth
def export_by_location(location_name):
    """Export resources by specific location."""
    try:
        format_type = request.args.get('format', 'excel').lower()
        
        if format_type == 'excel':
            return export_excel_with_filter({'location': location_name}, f"Location_{location_name}")
        else:
            return export_csv_with_filter({'location': location_name}, f"Location_{location_name}")
            
    except Exception as e:
        logger.error(f"Error exporting by location: {e}")
        return jsonify({'error': 'Failed to export by location'}), 500

@export_bp.route('/department/<department_name>', methods=['GET'])
@require_auth
def export_by_department(department_name):
    """Export resources by department."""
    try:
        format_type = request.args.get('format', 'excel').lower()
        
        if format_type == 'excel':
            return export_excel_with_filter({'department': department_name}, f"Department_{department_name}")
        else:
            return export_csv_with_filter({'department': department_name}, f"Department_{department_name}")
            
    except Exception as e:
        logger.error(f"Error exporting by department: {e}")
        return jsonify({'error': 'Failed to export by department'}), 500

@export_bp.route('/filtered', methods=['GET'])
@require_auth
def export_filtered():
    """Export filtered resources."""
    try:
        format_type = request.args.get('format', 'excel').lower()
        
        # Get filter parameters
        query_filter = {}
        
        department = request.args.get('department')
        location = request.args.get('location')
        device_type = request.args.get('device_type')
        search = request.args.get('search')
        
        if department:
            query_filter['department'] = department
        if location:
            query_filter['location'] = location
        if device_type:
            query_filter['device_name'] = {'$regex': device_type, '$options': 'i'}
        if search:
            query_filter['$or'] = [
                {'device_name': {'$regex': search, '$options': 'i'}},
                {'description': {'$regex': search, '$options': 'i'}},
                {'location': {'$regex': search, '$options': 'i'}}
            ]
        
        if format_type == 'excel':
            return export_excel_with_filter(query_filter, "Filtered_Resources")
        else:
            return export_csv_with_filter(query_filter, "Filtered_Resources")
            
    except Exception as e:
        logger.error(f"Error exporting filtered: {e}")
        return jsonify({'error': 'Failed to export filtered data'}), 500

# ============================================================================
# BULK EXPORT OPERATIONS
# ============================================================================

@export_bp.route('/bulk', methods=['GET'])
@require_auth
def bulk_export():
    """Generate bulk export package with multiple formats and reports."""
    try:
        db = get_db()
        
        # Create temporary directory for files
        temp_dir = tempfile.mkdtemp()
        
        # Get all resources
        resources = list(db.resources.find({}).sort('sl_no', 1))
        departments = db.resources.distinct('department')
        
        files_created = []
        
        # Generate comprehensive reports
        analytics = generate_inventory_analytics(resources)
        
        # 1. Full inventory Excel
        full_excel_path = os.path.join(temp_dir, 'full_inventory_report.xlsx')
        with open(full_excel_path, 'wb') as f:
            excel_data = create_full_inventory_excel(resources, analytics)
            f.write(excel_data.get_data())
        files_created.append(('full_inventory_report.xlsx', full_excel_path))
        
        # 2. Full inventory CSV
        full_csv_path = os.path.join(temp_dir, 'full_inventory_report.csv')
        with open(full_csv_path, 'w', encoding='utf-8') as f:
            csv_data = create_full_inventory_csv(resources, analytics)
            f.write(csv_data.get_data(as_text=True))
        files_created.append(('full_inventory_report.csv', full_csv_path))
        
        # 3. Department-wise reports
        for dept in departments:
            dept_resources = [r for r in resources if r['department'] == dept]
            if dept_resources:
                dept_analytics = generate_department_analytics(dept, dept_resources)
                
                # Department Excel
                dept_excel_path = os.path.join(temp_dir, f'{dept.replace(" ", "_")}_report.xlsx')
                with open(dept_excel_path, 'wb') as f:
                    dept_excel_data = generate_department_excel_report(dept, dept_resources, dept_analytics)
                    f.write(dept_excel_data.get_data())
                files_created.append((f'{dept.replace(" ", "_")}_report.xlsx', dept_excel_path))
        
        # 4. Critical items report
        critical_items = [r for r in resources if r.get('cost', 0) > 100000]
        if critical_items:
            critical_excel_path = os.path.join(temp_dir, 'critical_items_report.xlsx')
            with open(critical_excel_path, 'wb') as f:
                critical_excel_data = create_critical_items_excel(critical_items, 100000)
                f.write(critical_excel_data.get_data())
            files_created.append(('critical_items_report.xlsx', critical_excel_path))
        
        # Create ZIP file
        zip_path = os.path.join(temp_dir, 'campus_assets_bulk_export.zip')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for filename, filepath in files_created:
                zipf.write(filepath, filename)
        
        # Read ZIP file
        with open(zip_path, 'rb') as f:
            zip_data = f.read()
        
        # Clean up temporary files
        import shutil
        shutil.rmtree(temp_dir)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'campus_assets_bulk_export_{timestamp}.zip'
        
        return Response(
            zip_data,
            mimetype='application/zip',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
        
    except Exception as e:
        logger.error(f"Error creating bulk export: {e}")
        return jsonify({'error': 'Failed to create bulk export'}), 500

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def export_excel_with_filter(query_filter, report_name="Filtered_Resources"):
    """Helper function to export Excel with filter."""
    db = get_db()
    resources = list(db.resources.find(query_filter).sort('sl_no', 1))
    
    if not resources:
        return jsonify({'error': 'No resources found matching the criteria'}), 404
    
    export_data = prepare_resource_data(resources)
    summary_stats = generate_summary_statistics(resources)
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Main data sheet
        df = pd.DataFrame(export_data)
        df.to_excel(writer, sheet_name='Resources', index=False)
        
        # Summary sheet
        summary_data = {
            'Metric': [
                'Total Items',
                'Total Quantity',
                'Total Value (₹)',
                'Average Cost (₹)',
                'Unique Departments',
                'Unique Locations',
                'Unique Device Types'
            ],
            'Value': [
                summary_stats['total_items'],
                summary_stats['total_quantity'],
                f"₹{summary_stats['total_value']:,.2f}",
                f"₹{summary_stats['average_cost_per_item']:,.2f}",
                summary_stats['unique_departments'],
                summary_stats['unique_locations'],
                summary_stats['unique_device_types']
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Apply styling
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            apply_excel_styling(worksheet, f"{report_name} Report")
    
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{report_name.lower()}_{timestamp}.xlsx'
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

def export_csv_with_filter(query_filter, report_name="Filtered_Resources"):
    """Helper function to export CSV with filter."""
    db = get_db()
    resources = list(db.resources.find(query_filter).sort('sl_no', 1))
    
    if not resources:
        return jsonify({'error': 'No resources found matching the criteria'}), 404
    
    export_data = prepare_resource_data(resources)
    
    output = io.StringIO()
    
    # Add header with filter info
    output.write(f"# {report_name} Report\n")
    output.write(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    output.write(f"# Total Items: {len(resources)}\n")
    output.write("#\n")
    
    fieldnames = ['sl_no', 'device_name', 'quantity', 'description', 'procurement_date', 
                 'location', 'cost', 'total_value', 'department']
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    
    writer.writeheader()
    writer.writerows(export_data)
    
    csv_content = output.getvalue()
    output.close()
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{report_name.lower()}_{timestamp}.csv'
    
    return Response(
        csv_content,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

# ============================================================================
# ERROR HANDLERS
# ============================================================================

@export_bp.errorhandler(400)
def bad_request(error):
    """Handle bad request errors."""
    return jsonify({'error': 'Bad request'}), 400

@export_bp.errorhandler(404)
def not_found(error):
    """Handle not found errors."""
    return jsonify({'error': 'Export data not found'}), 404

@export_bp.errorhandler(500)
def internal_error(error):
    """Handle internal server errors."""
    return jsonify({'error': 'Internal server error'}), 500
# extra
@export_bp.route('/pdf', methods=['GET'])
@require_auth
def export_pdf():
    """Export resources to PDF format with professional formatting."""
    try:
        db = get_db()
        
        # Get filter parameters
        department = request.args.get('department')
        location = request.args.get('location')
        device_type = request.args.get('device_type')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Build query filter
        query_filter = {}
        if department:
            query_filter['department'] = department
        if location:
            query_filter['location'] = location
        if device_type:
            query_filter['device_name'] = {'$regex': device_type, '$options': 'i'}
        
        # Date range filter
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter['$gte'] = datetime.fromisoformat(start_date)
            if end_date:
                date_filter['$lte'] = datetime.fromisoformat(end_date)
            if date_filter:
                query_filter['procurement_date'] = date_filter
        
        # Get resources
        resources = list(db.resources.find(query_filter).sort('sl_no', 1))
        
        if not resources:
            return jsonify({'error': 'No data found for export'}), 404
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Get styles
        styles = getSampleStyleSheet()
        
        # Create content
        story = []
        
        # Title
        title = Paragraph("Campus Assets Management System", styles['Title'])
        story.append(title)
        
        subtitle = Paragraph("Resource Inventory Report", styles['Heading2'])
        story.append(subtitle)
        story.append(Spacer(1, 20))
        
        # Report information
        report_info_text = f"""
        <b>Generated On:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
        <b>Total Records:</b> {len(resources)}<br/>
        <b>Department Filter:</b> {department or 'All Departments'}<br/>
        <b>Location Filter:</b> {location or 'All Locations'}<br/>
        <b>Device Type Filter:</b> {device_type or 'All Device Types'}<br/>
        """
        
        if start_date or end_date:
            date_range = f"{start_date or 'Beginning'} to {end_date or 'Present'}"
            report_info_text += f"<b>Date Range:</b> {date_range}<br/>"
        
        report_info = Paragraph(report_info_text, styles['Normal'])
        story.append(report_info)
        story.append(Spacer(1, 30))
        
        # Calculate summary statistics
        total_quantity = sum(r.get('quantity', 0) for r in resources)
        total_value = sum(r.get('cost', 0) * r.get('quantity', 0) for r in resources)
        avg_cost = sum(r.get('cost', 0) for r in resources) / len(resources) if resources else 0
        
        # Summary section
        summary_text = f"""
        <b>Summary Statistics:</b><br/>
        • Total Items: {total_quantity:,}<br/>
        • Total Asset Value: ₹{total_value:,.2f}<br/>
        • Average Item Cost: ₹{avg_cost:,.2f}<br/>
        • Unique Devices: {len(set(r.get('device_name', '') for r in resources))}<br/>
        • Unique Locations: {len(set(r.get('location', '') for r in resources))}
        """
        
        summary = Paragraph(summary_text, styles['Normal'])
        story.append(summary)
        story.append(Spacer(1, 30))
        
        # Resources table header
        table_title = Paragraph("<b>Resource Details</b>", styles['Heading3'])
        story.append(table_title)
        story.append(Spacer(1, 10))
        
        # Create table data (limit to first 50 items for PDF readability)
        display_resources = resources[:50]
        headers = ['SL#', 'Device Name', 'Qty', 'Location', 'Cost (₹)', 'Department']
        table_data = [headers]
        
        for resource in display_resources:
            device_name = str(resource.get('device_name', ''))
            # Truncate long device names
            if len(device_name) > 25:
                device_name = device_name[:22] + '...'
            
            location = str(resource.get('location', ''))
            # Truncate long location names
            if len(location) > 20:
                location = location[:17] + '...'
            
            department_name = str(resource.get('department', ''))
            # Truncate long department names
            if len(department_name) > 15:
                department_name = department_name[:12] + '...'
            
            row = [
                str(resource.get('sl_no', '')),
                device_name,
                str(resource.get('quantity', 0)),
                location,
                f"{resource.get('cost', 0):,.0f}",
                department_name
            ]
            table_data.append(row)
        
        # Create table with styling
        table = Table(table_data, colWidths=[40, 120, 30, 100, 80, 80])
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightgrey, colors.white])
        ]))
        
        story.append(table)
        
        # Add note if there are more records
        if len(resources) > 50:
            note = Paragraph(f"<i>Note: Showing first 50 records out of {len(resources)} total. Download CSV/Excel for complete data.</i>", styles['Italic'])
            story.append(Spacer(1, 20))
            story.append(note)
        
        # Footer
        story.append(Spacer(1, 30))
        footer = Paragraph("<i>Generated by Campus Assets Management System</i>", styles['Normal'])
        story.append(footer)
        
        # Build PDF
        doc.build(story)
        
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_parts = ['campus_assets_report']
        if department:
            filename_parts.append(department.replace(' ', '_').replace('&', 'and'))
        if location:
            filename_parts.append(location.replace(' ', '_'))
        filename_parts.append(timestamp)
        filename = f"{'_'.join(filename_parts)}.pdf"
        
        # Create response
        response = make_response(pdf_content)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        logger.info(f"PDF export generated: {filename} with {len(resources)} records")
        return response
        
    except Exception as e:
        logger.error(f"Error generating PDF export: {e}")
        return jsonify({'error': f'Failed to generate PDF export: {str(e)}'}), 500
@export_bp.route('/json', methods=['GET'])
@require_auth
def export_json():
    """Export resources to JSON format for API consumption."""
    try:
        db = get_db()
        
        # Get filter parameters
        department = request.args.get('department')
        location = request.args.get('location')
        device_type = request.args.get('device_type')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        include_stats = request.args.get('include_stats', 'true').lower() == 'true'
        
        # Build query filter
        query_filter = {}
        if department:
            query_filter['department'] = department
        if location:
            query_filter['location'] = location
        if device_type:
            query_filter['device_name'] = {'$regex': device_type, '$options': 'i'}
        
        # Date range filter
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter['$gte'] = datetime.fromisoformat(start_date)
            if end_date:
                date_filter['$lte'] = datetime.fromisoformat(end_date)
            if date_filter:
                query_filter['procurement_date'] = date_filter
        
        # Get resources
        resources = list(db.resources.find(query_filter).sort('sl_no', 1))
        
        if not resources:
            return jsonify({'error': 'No data found for export'}), 404
        
        # Serialize resources for JSON export
        serialized_resources = []
        for resource in resources:
            serialized_resource = {}
            
            for key, value in resource.items():
                if isinstance(value, ObjectId):
                    serialized_resource[key] = str(value)
                elif isinstance(value, datetime):
                    serialized_resource[key] = value.isoformat()
                elif hasattr(value, 'isoformat'):  # Handle date objects
                    serialized_resource[key] = value.isoformat()
                else:
                    serialized_resource[key] = value
            
            # Add calculated fields
            serialized_resource['total_value'] = resource.get('cost', 0) * resource.get('quantity', 0)
            
            serialized_resources.append(serialized_resource)
        
        # Prepare export data structure
        export_data = {
            'export_metadata': {
                'generated_at': datetime.now().isoformat(),
                'export_type': 'resources',
                'format': 'json',
                'total_records': len(resources),
                'filters_applied': {
                    'department': department,
                    'location': location,
                    'device_type': device_type,
                    'start_date': start_date,
                    'end_date': end_date
                },
                'user_info': {
                    'user_id': str(request.current_user['_id']),
                    'user_email': request.current_user.get('email'),
                    'export_timestamp': datetime.now().isoformat()
                }
            },
            'resources': serialized_resources
        }
        
        # Add statistics if requested
        if include_stats:
            total_quantity = sum(r.get('quantity', 0) for r in resources)
            total_value = sum(r.get('cost', 0) * r.get('quantity', 0) for r in resources)
            unique_departments = len(set(r.get('department', '') for r in resources))
            unique_locations = len(set(r.get('location', '') for r in resources))
            unique_devices = len(set(r.get('device_name', '') for r in resources))
            avg_cost = sum(r.get('cost', 0) for r in resources) / len(resources) if resources else 0
            
            # Cost distribution
            cost_ranges = {
                'low_cost': len([r for r in resources if r.get('cost', 0) < 10000]),
                'medium_cost': len([r for r in resources if 10000 <= r.get('cost', 0) < 50000]),
                'high_cost': len([r for r in resources if r.get('cost', 0) >= 50000])
            }
            
            # Department distribution
            dept_stats = {}
            for resource in resources:
                dept = resource.get('department', 'Unknown')
                if dept not in dept_stats:
                    dept_stats[dept] = {'count': 0, 'total_value': 0}
                dept_stats[dept]['count'] += resource.get('quantity', 0)
                dept_stats[dept]['total_value'] += resource.get('cost', 0) * resource.get('quantity', 0)
            
            export_data['statistics'] = {
                'summary': {
                    'total_resources': len(resources),
                    'total_quantity': total_quantity,
                    'total_asset_value': total_value,
                    'average_cost_per_item': avg_cost,
                    'unique_departments': unique_departments,
                    'unique_locations': unique_locations,
                    'unique_device_types': unique_devices
                },
                'cost_distribution': cost_ranges,
                'department_breakdown': dept_stats,
                'value_metrics': {
                    'highest_value_item': max((r.get('cost', 0) for r in resources), default=0),
                    'lowest_value_item': min((r.get('cost', 0) for r in resources), default=0),
                    'median_cost': sorted([r.get('cost', 0) for r in resources])[len(resources)//2] if resources else 0
                }
            }
        
        # Convert to JSON string with proper formatting
        json_content = json.dumps(export_data, indent=2, ensure_ascii=False, default=str)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_parts = ['campus_assets_export']
        if department:
            filename_parts.append(department.replace(' ', '_').replace('&', 'and'))
        if location:
            filename_parts.append(location.replace(' ', '_'))
        filename_parts.append(timestamp)
        filename = f"{'_'.join(filename_parts)}.json"
        
        # Create response
        response = make_response(json_content)
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        logger.info(f"JSON export generated: {filename} with {len(resources)} records")
        return response
        
    except Exception as e:
        logger.error(f"Error generating JSON export: {e}")
        return jsonify({'error': f'Failed to generate JSON export: {str(e)}'}), 500
@export_bp.route('/csv/template', methods=['GET'])
@require_auth
def export_csv_template():
    """Download CSV template for resource import."""
    try:
        # Define template structure with sample data
        template_data = [
            {
                'device_name': 'Desktop Computer',
                'quantity': 10,
                'description': 'High-performance desktop computer for laboratory use',
                'procurement_date': '2024-01-15',
                'location': 'Computer Lab A-101',
                'cost': 45000.00,
                'department': 'Computer Science & Engineering'
            },
            {
                'device_name': 'Oscilloscope',
                'quantity': 5,
                'description': 'Digital storage oscilloscope for electronics lab',
                'procurement_date': '2024-02-20',
                'location': 'Electronics Lab B-201',
                'cost': 75000.00,
                'department': 'Electrical & Electronics Engineering'
            },
            {
                'device_name': 'Projector',
                'quantity': 2,
                'description': 'Digital projector for classroom presentations',
                'procurement_date': '2024-03-10',
                'location': 'Classroom C-301',
                'cost': 35000.00,
                'department': 'General'
            }
        ]
        
        # Create CSV content
        output = io.StringIO()
        
        if template_data:
            # Get field names from the first record
            fieldnames = template_data[0].keys()
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            
            # Write header
            writer.writeheader()
            
            # Write sample data rows
            writer.writerows(template_data)
        
        csv_content = output.getvalue()
        output.close()
        
        # Add helpful comments at the top
        comments = [
            "# Campus Assets Management System - Import Template",
            "# Instructions:",
            "# 1. Fill in your resource data below the header row",
            "# 2. Keep the column names exactly as shown",
            "# 3. Date format: YYYY-MM-DD (e.g., 2024-01-15)",
            "# 4. Cost should be in numbers only (e.g., 45000.00)",
            "# 5. Required fields: device_name, quantity, description, location, cost",
            "# 6. Optional fields: procurement_date, department",
            "# 7. Remove these comment lines before importing",
            "#",
            "# Sample data is provided below - replace with your actual data",
            "#"
        ]
        
        # Prepend comments to CSV content
        final_content = '\n'.join(comments) + '\n' + csv_content
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"campus_assets_import_template_{timestamp}.csv"
        
        # Create response
        response = make_response(final_content)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        logger.info(f"CSV template downloaded: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"Error generating CSV template: {e}")
        return jsonify({'error': f'Failed to generate CSV template: {str(e)}'}), 500
